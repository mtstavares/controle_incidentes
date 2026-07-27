import hashlib
import hmac
import json
import os
import re
import tempfile
import unicodedata
import uuid
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, time
from pathlib import Path

import pandas as pd
from flask import current_app
from openpyxl import load_workbook
from sqlalchemy import and_, or_
from werkzeug.utils import secure_filename

from app import db
from app.models import CredencialColetaMensal, CredencialComprometida, CredencialImportLote
from app.services.timezone_service import APP_TIMEZONE, utc_now


ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
MAX_SPREADSHEET_SIZE = 10 * 1024 * 1024
MAX_SEARCH_LENGTH = 80
ACCESS_FILTERS = {"", "somente_ad", "somente_ms", "ad_ms", "nenhum", "alguma_aplicacao"}
EMAIL_NOT_FOUND = "e-mail não localizado"
ORDER_FIELDS = {
    "data_coleta": CredencialComprometida.data_coleta,
    "nome": CredencialComprometida.nome_busca,
    "cpf": CredencialComprometida.cpf,
    "id": CredencialComprometida.id,
}
ORDER_DIRECTIONS = {"asc", "desc"}

REQUIRED_COLUMNS = {
    "nome": "NOME",
    "cpf": "CPF",
    "email": "EMAIL",
    "data_coleta": "DATA COLETA",
    "acesso_ad": "ACESSO AD",
    "acesso_ms": "ACESSO MS",
    "situacao_legal": "Situação legal",
    "mensagem_bloqueio": "MSG BLOQUEIO.",
}

OPTIONAL_COLUMNS = {
    "url": "URL",
    "permitiu_acesso": "Permitiu acesso a alguma aplicação?",
    "observacoes": "OBSERVAÇÕES",
}

MONTHLY_REQUIRED_SHEETS = ("Credenciais AD", "Credenciais MS", "Total")
MONTHLY_POSITIVE_SHEETS = ("Credenciais AD", "Credenciais MS")
MONTHLY_SHEET_REQUIRED_COLUMNS = {
    "nome",
    "cpf",
    "senha",
    "email",
    "url",
    "data_coleta",
    "quantidade_identificacoes",
    "data_identificacoes",
    "fonte",
    "acesso_ad",
    "acesso_ms",
    "situacao_legal",
}
MONTHLY_POSITIVE_REQUIRED_COLUMNS = {
    "observacoes",
    "mensagem_bloqueio_rds",
}
MONTHLY_PREVIEW_DIR = "credential_import_previews"
MONTHLY_BATCH_ACTIVE = "ativo"
MONTHLY_BATCH_REPLACED = "substituido"

COLUMN_ALIASES = {
    "nome": "nome",
    "cpf": "cpf",
    "email": "email",
    "url": "url",
    "data coleta": "data_coleta",
    "permitiu acesso a alguma aplicacao": "permitiu_acesso",
    "acesso ad": "acesso_ad",
    "acesso ms": "acesso_ms",
    "situacao legal": "situacao_legal",
    "situa o legal": "situacao_legal",
    "observacoes": "observacoes",
    "observa es": "observacoes",
    "msg bloqueio": "mensagem_bloqueio",
    "msg bloqueio.": "mensagem_bloqueio",
    "msg bloqueio rds": "mensagem_bloqueio_rds",
    "mensagem de bloqueio rds": "mensagem_bloqueio_rds",
    "mensagem bloqueio rds": "mensagem_bloqueio_rds",
    "mnesagem de bloqueio rds": "mensagem_bloqueio_rds",
    "mnesagem bloqueio rds": "mensagem_bloqueio_rds",
    "rds": "rds",
    "quantidade de identificacoes": "quantidade_identificacoes",
    "quantidade de identifica es": "quantidade_identificacoes",
    "data das identificacoes": "data_identificacoes",
    "data das identifica es": "data_identificacoes",
    "fonte": "fonte",
    "senha": "senha",
}

TRUE_VALUES = {"sim", "s", "true", "1", "positivo", "positiva", "yes", "y"}
FALSE_VALUES = {"nao", "não", "n", "false", "0", "negativo", "negativa", "no"}


@dataclass
class ImportSummary:
    total_rows: int = 0
    imported: int = 0
    updated: int = 0
    duplicates_ignored: int = 0
    rejected: int = 0
    ignored_password_column: bool = False
    errors: list[dict] = field(default_factory=list)
    positive_by_month: dict[int, int] = field(default_factory=dict)
    positive_by_competence: dict[str, int] = field(default_factory=dict)


MISSING_INFORMATION_TEXT = "Não foi possível encontrar informações"


@dataclass
class MonthlyImportPreview:
    token: str
    file_hash: str
    original_filename: str
    year: int
    month: int
    total_tested: int
    total_validated: int
    only_ad: int
    only_ms: int
    ad_and_ms: int
    not_validated: int
    positive_records: list[dict] = field(default_factory=list)
    errors: list[dict] = field(default_factory=list)
    warnings: list[dict] = field(default_factory=list)
    ignored_password_column: bool = False

    @property
    def can_confirm(self):
        return not self.errors

    def to_dict(self):
        return {
            "token": self.token,
            "file_hash": self.file_hash,
            "original_filename": self.original_filename,
            "year": self.year,
            "month": self.month,
            "total_tested": self.total_tested,
            "total_validated": self.total_validated,
            "only_ad": self.only_ad,
            "only_ms": self.only_ms,
            "ad_and_ms": self.ad_and_ms,
            "not_validated": self.not_validated,
            "positive_records": self.positive_records,
            "errors": self.errors,
            "warnings": self.warnings,
            "ignored_password_column": self.ignored_password_column,
        }

    @classmethod
    def from_dict(cls, payload):
        return cls(**payload)


def normalize_text(value, *, max_length=None, preserve_newlines=False):
    if value is None or pd.isna(value):
        return None
    text = str(value)
    text = text.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore")
    text = unicodedata.normalize("NFC", text)
    text = "".join(ch for ch in text if ch in "\r\n\t" or not unicodedata.category(ch).startswith("C"))
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    if preserve_newlines:
        text = "\n".join(re.sub(r"[ \t]+", " ", line).strip() for line in text.split("\n"))
        text = re.sub(r"\n{3,}", "\n\n", text)
    else:
        text = re.sub(r"\s+", " ", text).strip()
    if text[:1] in {"=", "+", "-", "@"}:
        text = "'" + text
    if max_length and len(text) > max_length:
        text = text[:max_length].rstrip()
    return text or None


def normalize_key(value):
    text = normalize_text(value) or ""
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip().casefold()


def normalize_column_name(value):
    return COLUMN_ALIASES.get(normalize_key(value), normalize_key(value).replace(" ", "_"))


def normalize_cpf(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 10:
        return "0" + digits
    return digits


def is_valid_cpf(cpf):
    if not re.fullmatch(r"\d{11}", cpf):
        return False
    if cpf == cpf[0] * 11:
        return False
    total = sum(int(cpf[i]) * (10 - i) for i in range(9))
    digit = (total * 10) % 11
    if digit == 10:
        digit = 0
    if digit != int(cpf[9]):
        return False
    total = sum(int(cpf[i]) * (11 - i) for i in range(10))
    digit = (total * 10) % 11
    if digit == 10:
        digit = 0
    return digit == int(cpf[10])


def format_cpf(cpf):
    digits = normalize_cpf(cpf)
    if len(digits) != 11:
        return digits
    return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"


def mask_cpf(cpf):
    digits = normalize_cpf(cpf)
    return f"***.***.***-{digits[-2:]}" if len(digits) >= 2 else "***.***.***-**"


def normalize_email(value):
    email = (normalize_text(value, max_length=255) or "").strip().lower()
    return email


def is_valid_email(email):
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email or ""))


def normalize_bool(value):
    key = normalize_key(value)
    if not key:
        return False
    if key in TRUE_VALUES:
        return True
    if key in FALSE_VALUES:
        return False
    return False


def normalize_bool_strict(value):
    key = normalize_key(value)
    if not key:
        return False, None
    if key in TRUE_VALUES:
        return True, None
    if key in FALSE_VALUES:
        return False, None
    return False, "valor de acesso ambiguo"


def parse_collection_date(value):
    if value is None or pd.isna(value):
        return None
    raw_value = str(value).strip()
    pt_months = {
        "JAN": "01",
        "FEV": "02",
        "MAR": "03",
        "ABR": "04",
        "MAI": "05",
        "JUN": "06",
        "JUL": "07",
        "AGO": "08",
        "SET": "09",
        "OUT": "10",
        "NOV": "11",
        "DEZ": "12",
    }
    compact_match = re.fullmatch(r"(\d{1,2})([A-Za-zÇÃÕÁÉÍÓÚÂÊÔ]{3})(\d{2,4})", raw_value.upper())
    if compact_match and compact_match.group(2) in pt_months:
        day, month_text, year = compact_match.groups()
        normalized_year = int(year)
        if normalized_year < 100:
            normalized_year += 2000
        raw_value = f"{int(day):02d}/{pt_months[month_text]}/{normalized_year:04d}"

    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw_value):
        parsed = pd.to_datetime(raw_value, errors="coerce", format="%Y-%m-%d")
    elif re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}", raw_value):
        parsed = pd.to_datetime(raw_value, errors="coerce", format="%Y-%m-%d %H:%M:%S")
    else:
        parsed = pd.to_datetime(raw_value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    dt = parsed.to_pydatetime()
    if dt.tzinfo is None:
        dt = datetime.combine(dt.date(), dt.time() if dt.time() != time.min else time.min, tzinfo=APP_TIMEZONE)
    else:
        dt = dt.astimezone(APP_TIMEZONE)
    return dt


def _read_spreadsheet(path):
    suffix = path.suffix.lower()
    engine = "openpyxl" if suffix == ".xlsx" else "xlrd"
    return pd.read_excel(path, dtype=str, engine=engine)


def validate_spreadsheet_file(storage):
    filename = secure_filename(storage.filename or "")
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("Envie uma planilha Excel no formato .xlsx ou .xls.")

    stream = storage.stream
    position = stream.tell()
    stream.seek(0, 2)
    size = stream.tell()
    stream.seek(position)
    if size <= 0:
        raise ValueError("A planilha enviada está vazia.")
    if size > MAX_SPREADSHEET_SIZE:
        raise ValueError("A planilha excede o tamanho máximo permitido.")
    if suffix == ".xlsx":
        signature = stream.read(4)
        stream.seek(position)
        if signature != b"PK\x03\x04":
            raise ValueError("A planilha enviada não possui assinatura XLSX válida.")
    return suffix


def _row_value(row, key):
    value = row.get(key)
    return None if value is None or pd.isna(value) else value


def _build_record(row):
    cpf = normalize_cpf(_row_value(row, "cpf"))
    email = normalize_email(_row_value(row, "email"))
    if not email:
        email = EMAIL_NOT_FOUND
    nome = normalize_text(_row_value(row, "nome"), max_length=255)
    data_coleta = parse_collection_date(_row_value(row, "data_coleta"))
    acesso_ad = normalize_bool(_row_value(row, "acesso_ad"))
    acesso_ms = normalize_bool(_row_value(row, "acesso_ms"))
    permitiu_acesso = normalize_bool(_row_value(row, "permitiu_acesso")) or acesso_ad or acesso_ms
    situacao_legal = normalize_text(_row_value(row, "situacao_legal"), max_length=150)

    errors = []
    if not nome:
        errors.append("nome ausente")
    if not is_valid_cpf(cpf):
        errors.append("CPF inválido")
    if email != EMAIL_NOT_FOUND and not is_valid_email(email):
        errors.append("e-mail inválido")
    if not data_coleta:
        errors.append("data de coleta inválida")
    if not situacao_legal:
        errors.append("situação legal ausente")

    return {
        "nome": nome,
        "nome_busca": normalize_key(nome),
        "cpf": cpf,
        "email": email,
        "url_origem": normalize_text(_row_value(row, "url"), max_length=2000),
        "data_coleta": data_coleta,
        "permitiu_acesso": permitiu_acesso,
        "acesso_ad": acesso_ad,
        "acesso_ms": acesso_ms,
        "situacao_legal": situacao_legal,
        "situacao_legal_normalizada": normalize_key(situacao_legal) if situacao_legal else None,
        "observacoes": normalize_text(_row_value(row, "observacoes"), max_length=4000, preserve_newlines=True),
        "mensagem_bloqueio": normalize_text(_row_value(row, "mensagem_bloqueio"), max_length=1000, preserve_newlines=True),
    }, errors


def _missing_to_default(value, *, max_length=None, preserve_newlines=False):
    return normalize_text(value, max_length=max_length, preserve_newlines=preserve_newlines) or MISSING_INFORMATION_TEXT


def _build_positive_year_record(row, allowed_years):
    cpf = normalize_cpf(_row_value(row, "cpf"))
    email = normalize_email(_row_value(row, "email")) or MISSING_INFORMATION_TEXT
    nome = _missing_to_default(_row_value(row, "nome"), max_length=255)
    data_coleta = parse_collection_date(_row_value(row, "data_coleta"))
    acesso_ad, acesso_ad_error = normalize_bool_strict(_row_value(row, "acesso_ad"))
    acesso_ms, acesso_ms_error = normalize_bool_strict(_row_value(row, "acesso_ms"))
    permitiu_acesso_raw, permitiu_acesso_error = normalize_bool_strict(_row_value(row, "permitiu_acesso"))
    permitiu_acesso = permitiu_acesso_raw or acesso_ad or acesso_ms
    situacao_legal = _missing_to_default(_row_value(row, "situacao_legal"), max_length=150)

    errors = []
    if not is_valid_cpf(cpf):
        errors.append("CPF invalido")
    if email != MISSING_INFORMATION_TEXT and not is_valid_email(email):
        errors.append("e-mail invalido")
    if not data_coleta:
        errors.append("data de coleta invalida")
    elif data_coleta.year not in allowed_years:
        errors.append("data de coleta fora do periodo permitido")
    elif data_coleta.year == 2024 and data_coleta.month == 1:
        errors.append("janeiro de 2024 nao possui carga cadastrada")
    if acesso_ad_error:
        errors.append("ACESSO AD ambiguo")
    if acesso_ms_error:
        errors.append("ACESSO MS ambiguo")
    if permitiu_acesso_error:
        errors.append("acesso geral ambiguo")
    if not permitiu_acesso:
        errors.append("sem acesso positivo confirmado")

    return {
        "nome": nome,
        "nome_busca": normalize_key(nome),
        "cpf": cpf,
        "email": email,
        "url_origem": _missing_to_default(_row_value(row, "url"), max_length=2000),
        "data_coleta": data_coleta,
        "permitiu_acesso": permitiu_acesso,
        "acesso_ad": acesso_ad,
        "acesso_ms": acesso_ms,
        "situacao_legal": situacao_legal,
        "situacao_legal_normalizada": normalize_key(situacao_legal),
        "observacoes": _missing_to_default(_row_value(row, "observacoes"), max_length=4000, preserve_newlines=True),
        "mensagem_bloqueio": _missing_to_default(
            _row_value(row, "mensagem_bloqueio"),
            max_length=1000,
            preserve_newlines=True,
        ),
    }, errors


def _find_existing(record):
    return CredencialComprometida.query.filter(
        CredencialComprometida.cpf == record["cpf"],
        CredencialComprometida.email == record["email"],
        CredencialComprometida.url_origem == record["url_origem"],
        CredencialComprometida.data_coleta == record["data_coleta"],
    ).first()


def _find_existing_positive(record):
    return CredencialComprometida.query.filter(
        CredencialComprometida.cpf == record["cpf"],
        CredencialComprometida.email == record["email"],
        CredencialComprometida.url_origem == record["url_origem"],
        db.func.date(CredencialComprometida.data_coleta) == record["data_coleta"].date().isoformat(),
    ).first()


def _merge_record(existing, record, user_id):
    changed = False
    for field, value in record.items():
        if value in (None, "") and getattr(existing, field) not in (None, ""):
            continue
        if getattr(existing, field) != value:
            setattr(existing, field, value)
            changed = True
    if changed:
        existing.imported_at = utc_now()
        existing.imported_by_id = user_id
    return changed


def _is_more_complete(value, current_value):
    if value in (None, ""):
        return False
    if current_value == value:
        return False
    if value == MISSING_INFORMATION_TEXT and current_value not in (None, "", MISSING_INFORMATION_TEXT, EMAIL_NOT_FOUND):
        return False
    return current_value in (None, "", MISSING_INFORMATION_TEXT, EMAIL_NOT_FOUND) or current_value != value


def _merge_positive_record(existing, record, user_id):
    changed = False
    for field, value in record.items():
        if field in {"cpf", "email", "url_origem", "data_coleta"}:
            continue
        if field in {"permitiu_acesso", "acesso_ad", "acesso_ms"}:
            new_value = bool(getattr(existing, field)) or bool(value)
            if getattr(existing, field) != new_value:
                setattr(existing, field, new_value)
                changed = True
            continue
        current_value = getattr(existing, field)
        if _is_more_complete(value, current_value):
            setattr(existing, field, value)
            changed = True
    if changed:
        existing.imported_at = utc_now()
        existing.imported_by_id = user_id
    return changed


def import_positive_credential_spreadsheet(storage, *, user_id=None, allowed_years=None):
    allowed_years = set(allowed_years or {2024})
    suffix = validate_spreadsheet_file(storage)
    summary = ImportSummary()
    temp_path = None
    temp_dir = Path(current_app.instance_path) / "tmp"
    temp_dir.mkdir(parents=True, exist_ok=True)

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix, dir=temp_dir) as temp_file:
            temp_path = Path(temp_file.name)
            storage.stream.seek(0)
            while True:
                chunk = storage.stream.read(1024 * 1024)
                if not chunk:
                    break
                temp_file.write(chunk)

        df = _read_spreadsheet(temp_path)
        df.columns = [normalize_column_name(column) for column in df.columns]
        if "senha" in df.columns:
            df = df.drop(columns=["senha"])
            summary.ignored_password_column = True

        missing = [label for key, label in REQUIRED_COLUMNS.items() if key not in df.columns]
        if missing:
            raise ValueError(f"Colunas obrigatórias ausentes: {', '.join(missing)}.")

        for key in OPTIONAL_COLUMNS:
            if key not in df.columns:
                df[key] = None

        seen_keys = set()
        summary.total_rows = int(len(df.index))
        for index, row in df.iterrows():
            line_number = int(index) + 2
            record, errors = _build_positive_year_record(row, allowed_years)
            if errors:
                summary.rejected += 1
                summary.errors.append({"linha": line_number, "campo": "validacao", "motivo": "; ".join(errors)})
                continue

            dedup_key = (
                record["cpf"],
                record["email"],
                record["url_origem"],
                record["data_coleta"].date().isoformat(),
            )
            if dedup_key in seen_keys:
                summary.duplicates_ignored += 1
                continue
            seen_keys.add(dedup_key)

            existing = _find_existing_positive(record)
            if existing:
                if _merge_positive_record(existing, record, user_id):
                    summary.updated += 1
                else:
                    summary.duplicates_ignored += 1
            else:
                db.session.add(CredencialComprometida(**record, imported_at=utc_now(), imported_by_id=user_id))
                summary.imported += 1

            month = record["data_coleta"].month
            competence = f"{record['data_coleta'].year:04d}-{month:02d}"
            summary.positive_by_month[month] = summary.positive_by_month.get(month, 0) + 1
            summary.positive_by_competence[competence] = summary.positive_by_competence.get(competence, 0) + 1

        return summary
    finally:
        if temp_path and temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                current_app.logger.warning("Não foi possível remover arquivo temporário de credenciais.")


def import_positive_2024_credential_spreadsheet(storage, user_id=None):
    return import_positive_credential_spreadsheet(storage, user_id=user_id, allowed_years={2024})


def import_credential_spreadsheet(storage, user_id=None):
    suffix = validate_spreadsheet_file(storage)
    summary = ImportSummary()
    temp_path = None
    temp_dir = Path(current_app.instance_path) / "tmp"
    temp_dir.mkdir(parents=True, exist_ok=True)

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix, dir=temp_dir) as temp_file:
            temp_path = Path(temp_file.name)
            storage.stream.seek(0)
            while True:
                chunk = storage.stream.read(1024 * 1024)
                if not chunk:
                    break
                temp_file.write(chunk)

        df = _read_spreadsheet(temp_path)
        df.columns = [normalize_column_name(column) for column in df.columns]
        if "senha" in df.columns:
            df = df.drop(columns=["senha"])
            summary.ignored_password_column = True

        missing = [label for key, label in REQUIRED_COLUMNS.items() if key not in df.columns]
        if missing:
            raise ValueError(f"Colunas obrigatórias ausentes: {', '.join(missing)}.")

        for key in OPTIONAL_COLUMNS:
            if key not in df.columns:
                df[key] = None

        summary.total_rows = int(len(df.index))
        for index, row in df.iterrows():
            line_number = int(index) + 2
            record, errors = _build_record(row)
            if errors:
                summary.rejected += 1
                summary.errors.append({"linha": line_number, "motivo": "; ".join(errors)})
                continue

            existing = _find_existing(record)
            if existing:
                if _merge_record(existing, record, user_id):
                    summary.updated += 1
                continue

            db.session.add(CredencialComprometida(**record, imported_at=utc_now(), imported_by_id=user_id))
            summary.imported += 1

        return summary
    finally:
        if temp_path and temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                current_app.logger.warning("Não foi possível remover arquivo temporário de credenciais.")


def validate_reference_year(year):
    if isinstance(year, bool):
        raise ValueError("Ano de referencia invalido.")
    try:
        year = int(year)
    except (TypeError, ValueError):
        raise ValueError("Ano de referencia invalido.")
    if year < 2000 or year > 2100:
        raise ValueError("Ano de referencia invalido.")
    return year


def _monthly_preview_dir():
    path = Path(current_app.instance_path) / "tmp" / MONTHLY_PREVIEW_DIR
    path.mkdir(parents=True, exist_ok=True)
    return path


def _monthly_preview_path(token):
    if not re.fullmatch(r"[a-f0-9]{32}", str(token or "")):
        raise ValueError("Token de importacao invalido.")
    return _monthly_preview_dir() / f"{token}.json"


def _save_monthly_preview(preview):
    path = _monthly_preview_path(preview.token)
    path.write_text(json.dumps(preview.to_dict(), ensure_ascii=False, default=str), encoding="utf-8")


def load_monthly_import_preview(token):
    path = _monthly_preview_path(token)
    if not path.exists():
        raise ValueError("Previa de importacao expirada ou inexistente.")
    return MonthlyImportPreview.from_dict(json.loads(path.read_text(encoding="utf-8")))


def delete_monthly_import_preview(token):
    try:
        _monthly_preview_path(token).unlink(missing_ok=True)
    except OSError:
        current_app.logger.warning("Nao foi possivel remover previa temporaria de credenciais.")


def _validate_monthly_xlsx(storage):
    filename = secure_filename(storage.filename or "")
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("Envie a planilha mensal no formato .xlsx.")
    stream = storage.stream
    position = stream.tell()
    stream.seek(0, os.SEEK_END)
    size = stream.tell()
    stream.seek(0)
    if size <= 0:
        raise ValueError("A planilha enviada esta vazia.")
    if size > MAX_SPREADSHEET_SIZE:
        raise ValueError("A planilha excede o tamanho maximo permitido.")
    signature = stream.read(4)
    stream.seek(position)
    if signature != b"PK\x03\x04":
        raise ValueError("A planilha enviada nao possui assinatura XLSX valida.")
    return filename


def _copy_upload_to_temp(storage):
    filename = _validate_monthly_xlsx(storage)
    temp_dir = Path(current_app.instance_path) / "tmp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=temp_dir) as temp_file:
        temp_path = Path(temp_file.name)
        storage.stream.seek(0)
        while True:
            chunk = storage.stream.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
            temp_file.write(chunk)
    return temp_path, filename, digest.hexdigest()


def _assert_xlsx_has_no_macros(path):
    try:
        with zipfile.ZipFile(path) as archive:
            names = {name.lower() for name in archive.namelist()}
    except zipfile.BadZipFile:
        raise ValueError("A planilha enviada esta corrompida.")
    if any(name.endswith("vbaproject.bin") for name in names):
        raise ValueError("Arquivos com macro nao sao permitidos.")


def _worksheet_header_map(ws):
    headers = {}
    duplicated = set()
    for col in range(1, ws.max_column + 1):
        normalized = normalize_column_name(ws.cell(1, col).value)
        if not normalized:
            continue
        if normalized in headers:
            duplicated.add(normalized)
        headers[normalized] = col
    return headers, duplicated


def _cell_by_key(ws, row, header_map, key):
    col = header_map.get(key)
    if not col:
        return None
    return ws.cell(row, col).value


def _iter_sheet_rows(ws, header_map):
    for row in range(2, ws.max_row + 1):
        values = [_cell_by_key(ws, row, header_map, key) for key in MONTHLY_SHEET_REQUIRED_COLUMNS]
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        yield row


def _row_has_formula_outside_password(ws, row, header_map):
    password_col = header_map.get("senha")
    for col in range(1, ws.max_column + 1):
        if col == password_col:
            continue
        value = ws.cell(row, col).value
        if isinstance(value, str) and value.startswith("="):
            return True
    return False


def _infer_competence_from_filename(filename):
    key = normalize_key(Path(filename or "").stem)
    month_aliases = {
        "jan": 1,
        "janeiro": 1,
        "fev": 2,
        "fevereiro": 2,
        "mar": 3,
        "marco": 3,
        "abr": 4,
        "abril": 4,
        "mai": 5,
        "maio": 5,
        "jun": 6,
        "junho": 6,
        "jul": 7,
        "julho": 7,
        "ago": 8,
        "agosto": 8,
        "set": 9,
        "setembro": 9,
        "out": 10,
        "outubro": 10,
        "nov": 11,
        "novembro": 11,
        "dez": 12,
        "dezembro": 12,
    }
    matches = []
    for alias, month in month_aliases.items():
        for match in re.finditer(rf"\b{alias}\s*(\d{{2}}|20\d{{2}})\b", key):
            year = int(match.group(1))
            matches.append((validate_reference_year(year + 2000 if year < 100 else year), month))
    unique = set(matches)
    if len(unique) != 1:
        raise ValueError("Nao foi possivel identificar a competencia pelo nome do arquivo. Use o padrao Credenciais_JUL26.xlsx.")
    return next(iter(unique))


def _normalize_month_date(value, year, month):
    if value is None or str(value).strip() == "":
        return None
    raw = str(value).strip()
    if re.fullmatch(r"\d{1,2}[A-Za-z]{3}", raw):
        raw = f"{raw}{str(year)[-2:]}"
    parsed = parse_collection_date(raw)
    if parsed and parsed.year == year and parsed.month == month:
        return parsed
    return None


def _fingerprint_for_credential(cpf, password, year, month):
    password_text = "" if password is None else str(password)
    message = f"{cpf}|{password_text}|{year:04d}-{month:02d}".encode("utf-8")
    secret = str(current_app.config.get("SECRET_KEY") or "").encode("utf-8")
    return hmac.new(secret, message, hashlib.sha256).hexdigest()


def _monthly_record_from_row(ws, row, header_map, year, month, sheet_name):
    cpf = normalize_cpf(_cell_by_key(ws, row, header_map, "cpf"))
    password = _cell_by_key(ws, row, header_map, "senha")
    data_coleta = _normalize_month_date(_cell_by_key(ws, row, header_map, "data_coleta"), year, month)
    acesso_ad, acesso_ad_error = normalize_bool_strict(_cell_by_key(ws, row, header_map, "acesso_ad"))
    acesso_ms, acesso_ms_error = normalize_bool_strict(_cell_by_key(ws, row, header_map, "acesso_ms"))
    errors = []
    if not is_valid_cpf(cpf):
        errors.append("CPF invalido")
    if password is None or str(password).strip() == "":
        errors.append("senha ausente para deduplicacao em memoria")
    if not data_coleta:
        errors.append("data de coleta fora da competencia ou invalida")
    if acesso_ad_error:
        errors.append("ACESSO AD ambiguo")
    if acesso_ms_error:
        errors.append("ACESSO MS ambiguo")

    fingerprint = _fingerprint_for_credential(cpf, password, year, month) if cpf and password else ""
    email = normalize_email(_cell_by_key(ws, row, header_map, "email")) or EMAIL_NOT_FOUND
    if email != EMAIL_NOT_FOUND and not is_valid_email(email):
        email = EMAIL_NOT_FOUND
    mensagem_bloqueio_rds = _missing_to_default(
        _cell_by_key(ws, row, header_map, "mensagem_bloqueio_rds"),
        max_length=1000,
        preserve_newlines=True,
    )
    record = {
        "source_sheet": sheet_name,
        "source_line": row,
        "fingerprint": fingerprint,
        "nome": _missing_to_default(_cell_by_key(ws, row, header_map, "nome"), max_length=255),
        "nome_busca": "",
        "cpf": cpf,
        "email": email,
        "url_origem": _missing_to_default(_cell_by_key(ws, row, header_map, "url"), max_length=2000),
        "data_coleta": data_coleta.isoformat() if data_coleta else None,
        "permitiu_acesso": acesso_ad or acesso_ms,
        "acesso_ad": acesso_ad,
        "acesso_ms": acesso_ms,
        "situacao_legal": _missing_to_default(_cell_by_key(ws, row, header_map, "situacao_legal"), max_length=150),
        "situacao_legal_normalizada": "",
        "observacoes": _missing_to_default(
            _cell_by_key(ws, row, header_map, "observacoes"),
            max_length=4000,
            preserve_newlines=True,
        ),
        "mensagem_bloqueio": mensagem_bloqueio_rds,
        "rds": None,
    }
    record["nome_busca"] = normalize_key(record["nome"])
    record["situacao_legal_normalizada"] = normalize_key(record["situacao_legal"])
    return record, errors


def _merge_monthly_positive(existing, incoming):
    existing["acesso_ad"] = bool(existing.get("acesso_ad")) or bool(incoming.get("acesso_ad"))
    existing["acesso_ms"] = bool(existing.get("acesso_ms")) or bool(incoming.get("acesso_ms"))
    existing["permitiu_acesso"] = existing["acesso_ad"] or existing["acesso_ms"]
    for field in (
        "nome",
        "nome_busca",
        "email",
        "url_origem",
        "data_coleta",
        "situacao_legal",
        "situacao_legal_normalizada",
        "observacoes",
        "mensagem_bloqueio",
        "rds",
    ):
        if _is_more_complete(incoming.get(field), existing.get(field)):
            existing[field] = incoming[field]
    existing.setdefault("source_lines", []).append({
        "sheet": incoming.get("source_sheet"),
        "line": incoming.get("source_line"),
    })


def build_monthly_import_preview(storage):
    temp_path = None
    try:
        temp_path, original_filename, file_hash = _copy_upload_to_temp(storage)
        year, month = _infer_competence_from_filename(original_filename)
        _assert_xlsx_has_no_macros(temp_path)
        workbook = load_workbook(temp_path, read_only=False, data_only=False)
        try:
            missing_sheets = [sheet for sheet in MONTHLY_REQUIRED_SHEETS if sheet not in workbook.sheetnames]
            if missing_sheets:
                raise ValueError(f"Abas obrigatorias ausentes: {', '.join(missing_sheets)}.")
            preview = MonthlyImportPreview(
                token=uuid.uuid4().hex,
                file_hash=file_hash,
                original_filename=original_filename,
                year=year,
                month=month,
                total_tested=0,
                total_validated=0,
                only_ad=0,
                only_ms=0,
                ad_and_ms=0,
                not_validated=0,
                ignored_password_column=True,
            )
            sheet_maps = {}
            for sheet_name in MONTHLY_REQUIRED_SHEETS:
                ws = workbook[sheet_name]
                if ws.merged_cells.ranges:
                    preview.errors.append({"aba": sheet_name, "linha": "-", "motivo": "celulas mescladas nao permitidas"})
                header_map, duplicated = _worksheet_header_map(ws)
                if duplicated:
                    preview.errors.append({"aba": sheet_name, "linha": 1, "motivo": "cabecalhos duplicados"})
                required_columns = set(MONTHLY_SHEET_REQUIRED_COLUMNS)
                if sheet_name in MONTHLY_POSITIVE_SHEETS:
                    required_columns |= MONTHLY_POSITIVE_REQUIRED_COLUMNS
                missing = sorted(required_columns - set(header_map))
                if missing:
                    preview.errors.append({"aba": sheet_name, "linha": 1, "motivo": "cabecalhos obrigatorios ausentes: " + ", ".join(missing)})
                sheet_maps[sheet_name] = header_map

            total_keys = set()
            total_ws = workbook["Total"]
            total_map = sheet_maps["Total"]
            for row in _iter_sheet_rows(total_ws, total_map):
                if _row_has_formula_outside_password(total_ws, row, total_map):
                    preview.errors.append({"aba": "Total", "linha": row, "motivo": "formula nao permitida"})
                    continue
                record, errors = _monthly_record_from_row(total_ws, row, total_map, year, month, "Total")
                if errors:
                    preview.errors.append({"aba": "Total", "linha": row, "motivo": "; ".join(errors)})
                    continue
                total_keys.add(record["fingerprint"])

            positives = {}
            for sheet_name in MONTHLY_POSITIVE_SHEETS:
                ws = workbook[sheet_name]
                header_map = sheet_maps[sheet_name]
                for row in _iter_sheet_rows(ws, header_map):
                    if _row_has_formula_outside_password(ws, row, header_map):
                        preview.errors.append({"aba": sheet_name, "linha": row, "motivo": "formula nao permitida"})
                        continue
                    record, errors = _monthly_record_from_row(ws, row, header_map, year, month, sheet_name)
                    expected_positive = record["acesso_ad"] if sheet_name.endswith("AD") else record["acesso_ms"]
                    if not expected_positive:
                        errors.append("linha da aba positiva sem acesso confirmado")
                    if errors:
                        preview.errors.append({"aba": sheet_name, "linha": row, "motivo": "; ".join(errors)})
                        continue
                    if record["fingerprint"] not in total_keys:
                        preview.warnings.append({"aba": sheet_name, "linha": row, "motivo": "credencial positiva ausente na aba Total"})
                    existing = positives.get(record["fingerprint"])
                    if existing:
                        _merge_monthly_positive(existing, record)
                    else:
                        record["source_lines"] = [{"sheet": sheet_name, "line": row}]
                        positives[record["fingerprint"]] = record
                    if record.get("mensagem_bloqueio") == MISSING_INFORMATION_TEXT:
                        preview.errors.append({"aba": sheet_name, "linha": row, "motivo": "MSG BLOQUEIO - RDS ausente"})

            preview.total_tested = len(total_keys)
            preview.positive_records = list(positives.values())
            preview.total_validated = len(preview.positive_records)
            preview.only_ad = sum(1 for item in preview.positive_records if item["acesso_ad"] and not item["acesso_ms"])
            preview.only_ms = sum(1 for item in preview.positive_records if item["acesso_ms"] and not item["acesso_ad"])
            preview.ad_and_ms = sum(1 for item in preview.positive_records if item["acesso_ad"] and item["acesso_ms"])
            preview.not_validated = max(preview.total_tested - preview.total_validated, 0)
            _save_monthly_preview(preview)
            return preview
        finally:
            workbook.close()
    finally:
        if temp_path and temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                current_app.logger.warning("Nao foi possivel remover arquivo temporario de credenciais.")


def _active_batch_for_competence(year, month):
    return CredencialImportLote.query.filter_by(
        ano_referencia=year,
        mes_referencia=month,
        status=MONTHLY_BATCH_ACTIVE,
    ).one_or_none()


def confirm_monthly_import(token, *, user_id=None):
    preview = load_monthly_import_preview(token)
    if not preview.can_confirm:
        raise ValueError("A previa possui erros e nao pode ser importada.")
    existing_hash = CredencialImportLote.query.filter_by(arquivo_sha256=preview.file_hash).one_or_none()
    if existing_hash:
        raise ValueError("Esta planilha ja foi importada anteriormente.")

    now = utc_now()
    previous = _active_batch_for_competence(preview.year, preview.month)
    last_version = (
        db.session.query(db.func.max(CredencialImportLote.versao))
        .filter_by(ano_referencia=preview.year, mes_referencia=preview.month)
        .scalar()
        or 0
    )
    batch = CredencialImportLote(
        arquivo_nome_original=preview.original_filename,
        arquivo_sha256=preview.file_hash,
        ano_referencia=preview.year,
        mes_referencia=preview.month,
        imported_at=now,
        imported_by_id=user_id,
        total_testado=preview.total_tested,
        total_validado=preview.total_validated,
        total_somente_ad=preview.only_ad,
        total_somente_ms=preview.only_ms,
        total_ad_ms=preview.ad_and_ms,
        total_nao_validado=preview.not_validated,
        rejeitados=len(preview.errors),
        status=MONTHLY_BATCH_ACTIVE,
        versao=last_version + 1,
        lote_substituido_id=previous.id if previous else None,
    )
    db.session.add(batch)
    db.session.flush()

    if previous:
        previous.status = MONTHLY_BATCH_REPLACED
        previous.deleted_at = now
        CredencialComprometida.query.filter_by(lote_id=previous.id).update({"deleted_at": now})

    for item in preview.positive_records:
        data_coleta = datetime.fromisoformat(item["data_coleta"])
        db.session.add(CredencialComprometida(
            nome=item["nome"],
            nome_busca=item["nome_busca"],
            cpf=item["cpf"],
            email=item["email"],
            url_origem=item["url_origem"],
            data_coleta=data_coleta,
            permitiu_acesso=item["permitiu_acesso"],
            acesso_ad=item["acesso_ad"],
            acesso_ms=item["acesso_ms"],
            situacao_legal=item["situacao_legal"],
            situacao_legal_normalizada=item["situacao_legal_normalizada"],
            observacoes=item["observacoes"],
            mensagem_bloqueio=item["mensagem_bloqueio"],
            rds=item.get("rds"),
            credencial_fingerprint=item["fingerprint"],
            lote_id=batch.id,
            imported_at=now,
            imported_by_id=user_id,
        ))

    monthly_total = CredencialColetaMensal.query.filter_by(
        ano_referencia=preview.year,
        mes_referencia=preview.month,
    ).one_or_none()
    if monthly_total is None:
        db.session.add(CredencialColetaMensal(
            ano_referencia=preview.year,
            mes_referencia=preview.month,
            quantidade_localizada=preview.total_tested,
        ))
    else:
        monthly_total.quantidade_localizada = preview.total_tested

    db.session.commit()
    delete_monthly_import_preview(token)
    return batch


def apply_credential_filters(query, args):
    search = (args.get("q") or "").strip()[:MAX_SEARCH_LENGTH]
    start_date = parse_collection_date(args.get("start_date"))
    end_date = parse_collection_date(args.get("end_date"))
    access_filter = args.get("access", "")
    situation = (args.get("situacao") or "").strip()

    if args.get("start_date") and not start_date:
        raise ValueError("Data inicial inválida.")
    if args.get("end_date") and not end_date:
        raise ValueError("Data final inválida.")
    if start_date and end_date and start_date.date() > end_date.date():
        raise ValueError("A data inicial não pode ser posterior à data final.")
    if access_filter not in ACCESS_FILTERS:
        raise ValueError("Filtro de acesso inválido.")

    if search:
        cpf_search = normalize_cpf(search)
        name_search = normalize_key(search)
        conditions = []
        if cpf_search:
            conditions.append(CredencialComprometida.cpf.like(f"%{cpf_search}%"))
        if name_search:
            conditions.append(CredencialComprometida.nome_busca.like(f"%{name_search}%"))
        if conditions:
            query = query.filter(or_(*conditions))

    if start_date:
        query = query.filter(CredencialComprometida.data_coleta >= datetime.combine(start_date.date(), time.min, tzinfo=APP_TIMEZONE))
    if end_date:
        query = query.filter(CredencialComprometida.data_coleta <= datetime.combine(end_date.date(), time.max, tzinfo=APP_TIMEZONE))

    if access_filter == "somente_ad":
        query = query.filter(and_(CredencialComprometida.acesso_ad.is_(True), CredencialComprometida.acesso_ms.is_(False)))
    elif access_filter == "somente_ms":
        query = query.filter(and_(CredencialComprometida.acesso_ad.is_(False), CredencialComprometida.acesso_ms.is_(True)))
    elif access_filter == "ad_ms":
        query = query.filter(and_(CredencialComprometida.acesso_ad.is_(True), CredencialComprometida.acesso_ms.is_(True)))
    elif access_filter == "nenhum":
        query = query.filter(and_(CredencialComprometida.acesso_ad.is_(False), CredencialComprometida.acesso_ms.is_(False)))
    elif access_filter == "alguma_aplicacao":
        query = query.filter(or_(
            CredencialComprometida.acesso_ad.is_(True),
            CredencialComprometida.acesso_ms.is_(True),
            CredencialComprometida.permitiu_acesso.is_(True),
        ))

    if situation:
        query = query.filter(CredencialComprometida.situacao_legal_normalizada == normalize_key(situation))

    return query


def order_credentials(query, args):
    field = args.get("sort", "data_coleta")
    direction = args.get("direction", "desc")
    if field not in ORDER_FIELDS:
        field = "data_coleta"
    if direction not in ORDER_DIRECTIONS:
        direction = "desc"
    column = ORDER_FIELDS[field]
    ordered = column.asc() if direction == "asc" else column.desc()
    return query.order_by(ordered, CredencialComprometida.id.desc()), field, direction


def credential_to_table_dict(item):
    data_coleta = item.data_coleta.strftime("%d/%m/%Y") if item.data_coleta else ""
    if item.acesso_ad and item.acesso_ms:
        sistema = "AD/MS"
    elif item.acesso_ad:
        sistema = "AD"
    elif item.acesso_ms:
        sistema = "MS"
    else:
        sistema = ""
    return {
        "id": item.id,
        "cpf": format_cpf(item.cpf),
        "data_coleta": data_coleta,
        "nome": item.nome,
        "email": item.email,
        "sistema": sistema,
        "rds": item.rds or "",
        "mensagem_bloqueio": item.mensagem_bloqueio or "",
        "situacao_legal": item.situacao_legal or "",
    }
